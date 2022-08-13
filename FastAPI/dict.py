import openpyxl

workbook = openpyxl.load_workbook('Menu.xlsx')
sheet = workbook['Menu']
name = sheet.cell(column=1, row=2).value
img = sheet.cell(column=2, row=2).value
info = sheet.cell(column=3, row=2).value
price = sheet.cell(column=4, row=2).value


menu_info = {}
menu_info['name'] = name
menu_info['img'] = img
menu_info['info'] = info
menu_info['price'] = price
