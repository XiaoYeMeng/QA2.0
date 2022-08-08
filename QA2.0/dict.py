import openpyxl
import random

# 讀取excel sheet
xlsx = openpyxl.load_workbook('./dict/dict.xlsx')
sheet = xlsx['dict']
# 第一題
topic_1_que = sheet.cell(column=1, row=1).value
topic_1_ans = sheet.cell(column=2, row=1).value
topic_1_a = sheet.cell(column=3, row=1).value
topic_1_b = sheet.cell(column=4, row=1).value
topic_1_c = sheet.cell(column=5, row=1).value
topic_1_d = sheet.cell(column=6, row=1).value
# 第二題
topic_2_que = sheet.cell(column=1, row=2).value
topic_2_ans = sheet.cell(column=2, row=2).value
topic_2_a = sheet.cell(column=3, row=2).value
topic_2_b = sheet.cell(column=4, row=2).value
topic_2_c = sheet.cell(column=5, row=2).value
topic_2_d = sheet.cell(column=6, row=2).value
# 第三題
topic_3_que = sheet.cell(column=1, row=3).value
topic_3_ans = sheet.cell(column=2, row=3).value
topic_3_a = sheet.cell(column=3, row=3).value
topic_3_b = sheet.cell(column=4, row=3).value
topic_3_c = sheet.cell(column=5, row=3).value
topic_3_d = sheet.cell(column=6, row=3).value
# 第四題
topic_4_que = sheet.cell(column=1, row=4).value
topic_4_ans = sheet.cell(column=2, row=4).value
topic_4_a = sheet.cell(column=3, row=4).value
topic_4_b = sheet.cell(column=4, row=4).value
topic_4_c = sheet.cell(column=5, row=4).value
topic_4_d = sheet.cell(column=6, row=4).value
# 第五題
topic_5_que = sheet.cell(column=1, row=5).value
topic_5_ans = sheet.cell(column=2, row=5).value
topic_5_a = sheet.cell(column=3, row=5).value
topic_5_b = sheet.cell(column=4, row=5).value
topic_5_c = sheet.cell(column=5, row=5).value
topic_5_d = sheet.cell(column=6, row=5).value
# 第六題
topic_6_que = sheet.cell(column=1, row=6).value
topic_6_ans = sheet.cell(column=2, row=6).value
topic_6_a = sheet.cell(column=3, row=6).value
topic_6_b = sheet.cell(column=4, row=6).value
topic_6_c = sheet.cell(column=5, row=6).value
topic_6_d = sheet.cell(column=6, row=6).value
# 第七題
topic_7_que = sheet.cell(column=1, row=7).value
topic_7_ans = sheet.cell(column=2, row=7).value
topic_7_a = sheet.cell(column=3, row=7).value
topic_7_b = sheet.cell(column=4, row=7).value
topic_7_c = sheet.cell(column=5, row=7).value
topic_7_d = sheet.cell(column=6, row=7).value
# 第八題
topic_8_que = sheet.cell(column=1, row=8).value
topic_8_ans = sheet.cell(column=2, row=8).value
topic_8_a = sheet.cell(column=3, row=8).value
topic_8_b = sheet.cell(column=4, row=8).value
topic_8_c = sheet.cell(column=5, row=8).value
topic_8_d = sheet.cell(column=6, row=8).value
# 第九題
topic_9_que = sheet.cell(column=1, row=9).value
topic_9_ans = sheet.cell(column=2, row=9).value
topic_9_a = sheet.cell(column=3, row=9).value
topic_9_b = sheet.cell(column=4, row=9).value
topic_9_c = sheet.cell(column=5, row=9).value
topic_9_d = sheet.cell(column=6, row=9).value
# 第十題
topic_10_que = sheet.cell(column=1, row=10).value
topic_10_ans = sheet.cell(column=2, row=10).value
topic_10_a = sheet.cell(column=3, row=10).value
topic_10_b = sheet.cell(column=4, row=10).value
topic_10_c = sheet.cell(column=5, row=10).value
topic_10_d = sheet.cell(column=6, row=10).value
# 第十一題
topic_11_que = sheet.cell(column=1, row=11).value
topic_11_ans = sheet.cell(column=2, row=11).value
topic_11_a = sheet.cell(column=3, row=11).value
topic_11_b = sheet.cell(column=4, row=11).value
topic_11_c = sheet.cell(column=5, row=11).value
topic_11_d = sheet.cell(column=6, row=11).value
# 第十二題
topic_12_que = sheet.cell(column=1, row=12).value
topic_12_ans = sheet.cell(column=2, row=12).value
topic_12_a = sheet.cell(column=3, row=12).value
topic_12_b = sheet.cell(column=4, row=12).value
topic_12_c = sheet.cell(column=5, row=12).value
topic_12_d = sheet.cell(column=6, row=12).value


topic = []
topic.append({'topic_que': topic_1_que, 'topic_ans': topic_1_ans, 'topic_a': topic_1_a,
             'topic_b': topic_1_b, 'topic_c': topic_1_c, 'topic_d': topic_1_d})
topic.append({'topic_que': topic_2_que, 'topic_ans': topic_2_ans, 'topic_a': topic_2_a,
             'topic_b': topic_2_b, 'topic_c': topic_2_c, 'topic_d': topic_2_d})
topic.append({'topic_que': topic_3_que, 'topic_ans': topic_3_ans, 'topic_a': topic_3_a,
             'topic_b': topic_3_b, 'topic_c': topic_3_c, 'topic_d': topic_3_d})
topic.append({'topic_que': topic_4_que, 'topic_ans': topic_4_ans, 'topic_a': topic_4_a,
             'topic_b': topic_4_b, 'topic_c': topic_4_c, 'topic_d': topic_4_d})
topic.append({'topic_que': topic_5_que, 'topic_ans': topic_5_ans, 'topic_a': topic_5_a,
             'topic_b': topic_5_b, 'topic_c': topic_5_c, 'topic_d': topic_5_d})
topic.append({'topic_que': topic_6_que, 'topic_ans': topic_6_ans, 'topic_a': topic_6_a,
             'topic_b': topic_6_b, 'topic_c': topic_6_c, 'topic_d': topic_6_d})
topic.append({'topic_que': topic_7_que, 'topic_ans': topic_7_ans, 'topic_a': topic_7_a,
             'topic_b': topic_7_b, 'topic_c': topic_7_c, 'topic_d': topic_7_d})
topic.append({'topic_que': topic_8_que, 'topic_ans': topic_8_ans, 'topic_a': topic_8_a,
             'topic_b': topic_8_b, 'topic_c': topic_8_c, 'topic_d': topic_8_d})
topic.append({'topic_que': topic_9_que, 'topic_ans': topic_9_ans, 'topic_a': topic_9_a,
             'topic_b': topic_9_b, 'topic_c': topic_9_c, 'topic_d': topic_9_d})
topic.append({'topic_que': topic_10_que, 'topic_ans': topic_10_ans, 'topic_a': topic_10_a,
             'topic_b': topic_10_b, 'topic_c': topic_10_c, 'topic_d': topic_10_d})
topic.append({'topic_que': topic_11_que, 'topic_ans': topic_11_ans, 'topic_a': topic_11_a,
             'topic_b': topic_11_b, 'topic_c': topic_11_c, 'topic_d': topic_11_d})
topic.append({'topic_que': topic_12_que, 'topic_ans': topic_12_ans, 'topic_a': topic_12_a,
             'topic_b': topic_12_b, 'topic_c': topic_12_c, 'topic_d': topic_12_d})

topic_random = random.sample(topic, 10)

# print(topic)
print(topic_random)


# for i in range(0, 10):
#     print(topic_random[i])

